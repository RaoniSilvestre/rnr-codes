import os
from openpyxl import Workbook, load_workbook


planilhas_filtradas = os.listdir("./Planilhas-filtradas/")

planilha_de_stats = "./Planilhas-de-output/Planilha-de-stats.xslx"

wb = Workbook()

assert wb.active is not None

ws = wb.active
ws.append(["Listas","Participantes totais", "Presentes", "Ausentes", "Presença"])

wb_presenças_por_pessoa = load_workbook("./Planilhas-de-output/Planilha-de-Participação.xlsx")

assert wb_presenças_por_pessoa.active is not None
presenças = wb_presenças_por_pessoa.active

total = 0
sims = 0

print(presenças.max_column)
for coluna in range(4, presenças.max_column - 1):
    for linha in range(2, presenças.max_row + 1):
        if presenças.cell(row=linha, column=coluna).value == "Sim":
            sims = sims + 1
        total = total + 1
    ws.cell(row=coluna-2, column=1).value = f"Lista{(coluna-3)}"   
    ws.cell(row=coluna-2, column=2).value = total
    ws.cell(row=coluna-2, column=3).value = sims
    ws.cell(row=coluna-2, column=4).value = total - sims
    ws.cell(row=coluna-2, column=5).value = f"{(sims*100/total):0.2f}%"

    print("\ntotal: ",total,"\nqntd de sim:", sims,f"\nporcentagem: {(sims*100/total):0.2f}",)
    total = 0
    sims = 0

total = 0
certificados = 0
for pessoa in range(2,presenças.max_row+1):
    if presenças.cell(row=pessoa, column=presenças.max_column).value == "RECEBE":
        certificados += 1
    total += 1
ws.cell(row=10, column=8).value = "Participantes com certificado"
ws.cell(row=11,column=8).value = f"{(certificados*100 / total):0.2f}%"

wb.save("./Planilhas-de-output/Planilha-de-stats.xlsx")




