import pandas as pd
from openpyxl import Workbook, load_workbook
import argparse

def criar_lista_filtrada(num):
    nome_da_planilha = "./Planilhas-base/Planilha-" + str(num) + ".xlsx"

    print("Planilha lida  : " + nome_da_planilha)
    df_original = pd.read_excel(nome_da_planilha, header=7)
    df_filtrado = df_original

    colunas_desejadas = ["Nome", "Sobrenome", "Email", "Check-in", "Data Check-in (*)", "Ramo"]

    df_filtrado_colunas_certas = df_filtrado[colunas_desejadas]

    wb = Workbook()
    
    assert wb.active is not None

    ws = wb.active

    headers = list(df_filtrado_colunas_certas.columns)

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    for _, row in df_filtrado_colunas_certas.iterrows():
        ws.append(row.tolist())

    nome_da_planilha_filtrada = "./Planilhas-filtradas/Planilha-"+ str(num) + "-filtrada.xlsx"
    wb.save(nome_da_planilha_filtrada)
    print("Planilha criada: " + nome_da_planilha_filtrada)


def criar_planilha_de_participação(num):
    print("Processando planilha de participação")
    wb = Workbook()
    assert wb.active is not None
    ws_new = wb.active
    headers = ["Nome","Sobrenome","Email"]
    for i in range(1,num+1):
        headers.append("Lista"+str(i))
    headers.append("Porcentagem")
    headers.append("Certificado")
    ws_new.append(headers)

    planilha_existente = load_workbook("./Planilhas-filtradas/Planilha-" + "1" + "-filtrada.xlsx")
    ws_existente = planilha_existente.active
    
    assert ws_existente is not None
   
    print("Inserindo Nome, Sobrenome, Email...")
    for linha in range(2, ws_existente.max_row + 1):    
        valor_nome = ws_existente.cell(row=linha, column=1).value
        valor_sobrenome = ws_existente.cell(row=linha, column=2).value
        valor_email = ws_existente.cell(row=linha, column=3).value  

        ws_new.cell(row=linha, column=1).value = valor_nome
        ws_new.cell(row=linha, column=2).value = valor_sobrenome
        ws_new.cell(row=linha, column=3).value = valor_email  
    print("Nome, Sobrenome, Email : OK")
     
    for k in range(1,num+1):
        print("Populando Check-in's da lista " + str(k))
        planilha_n = load_workbook("./Planilhas-filtradas/Planilha-"+ str(k) + "-filtrada.xlsx")
        ws_n = planilha_n.active
        
        assert ws_n is not None

        for linha in range(2, ws_n.max_row + 1):
            valor_checkin_n = ws_n.cell(row=linha, column=4)
            ws_new.cell(row=linha, column=k+3).value = valor_checkin_n.value
        print("Lista " + str(k) + " : OK")
    
    print("Verificando quem receberá presença")
    sims = 0
    total = 0
    total_listas = ws_new.max_row + 1
    print(total_listas)
    for pessoa in range(2, total_listas - 6):
        nome_sobrenome = str(ws_new.cell(row=pessoa, column=1).value) +" "+ str(ws_new.cell(row=pessoa, column=2).value)
        for presença in range(4, num+4):
            if ws_new.cell(row=pessoa, column=presença).value == "Sim":
                sims += 1
            total += 1
        
        presença = (sims*100/total)
        ws_new.cell(row=pessoa, column=(num+4)).value = f'{presença:0.2f}%'
        recebe_certificado = "RECEBE" if presença >= 75 else "NÃO RECEBE"
        ws_new.cell(row=pessoa, column=(num+5)).value = recebe_certificado 
        print(f"{nome_sobrenome} : {recebe_certificado}")
        total = 0
        sims = 0
    print("Verificação completa")
    wb.save("./Planilhas-de-output/Planilha-de-Participação.xlsx")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('quantidade', type=int, help="quantidade de listas de arquivos do tipo ./planilhas-base/planilha-N.xlsx")
    args = parser.parse_args()
    quantidade_listas = args.quantidade
    for i in range(1,quantidade_listas + 1):
        criar_lista_filtrada(i)
    criar_planilha_de_participação(quantidade_listas)

if __name__ == "__main__":
    main()

